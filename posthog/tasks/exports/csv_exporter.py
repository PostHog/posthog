import os
import csv
import json
import datetime
import tempfile
from collections import OrderedDict
from collections.abc import Generator, Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from typing import Any, Optional, Protocol
from urllib.parse import parse_qsl, quote, urlencode, urlparse, urlunparse

from django.http import QueryDict

import requests
import structlog
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from pydantic import BaseModel
from requests.exceptions import HTTPError
from rest_framework_csv.renderers import CSVRenderer

from posthog.api.services.query import process_query_dict
from posthog.exceptions_capture import capture_exception
from posthog.hogql_queries.query_runner import ExecutionMode
from posthog.jwt import PosthogJwtAudience, encode_jwt
from posthog.models.exported_asset import ExportedAsset, save_content_from_file
from posthog.utils import absolute_uri

from ...exceptions import QuerySizeExceeded
from ...hogql.constants import CSV_EXPORT_BREAKDOWN_LIMIT_INITIAL, CSV_EXPORT_BREAKDOWN_LIMIT_LOW, CSV_EXPORT_LIMIT
from ...hogql.query import LimitContext
from ...hogql_queries.insights.trends.breakdown import (
    BREAKDOWN_NULL_DISPLAY,
    BREAKDOWN_NULL_STRING_LABEL,
    BREAKDOWN_OTHER_DISPLAY,
    BREAKDOWN_OTHER_STRING_LABEL,
)
from ..exporter import EXPORT_TIMER

logger = structlog.get_logger(__name__)

RESULT_LIMIT_KEYS = ("distinct_ids",)
RESULT_LIMIT_LENGTH = 10
QUERY_PAGE_SIZE = 10000


def sanitize_value_for_excel(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)


class TabularWriter(Protocol):
    def write_header(self, columns: list[str]) -> None: ...
    def write_row(self, row: dict) -> None: ...
    def finish(self) -> str: ...  # returns path to temp file


class CsvWriter(TabularWriter):
    def __init__(self) -> None:
        self._tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
        self._writer: csv.DictWriter | None = None

    def write_header(self, columns: list[str]) -> None:
        self._writer = csv.DictWriter(self._tmp, fieldnames=columns, extrasaction="ignore")
        self._writer.writeheader()

    def write_row(self, row: dict) -> None:
        assert self._writer is not None
        self._writer.writerow(row)

    def finish(self) -> str:
        self._tmp.close()
        return self._tmp.name


class ExcelWriter(TabularWriter):
    def __init__(self) -> None:
        self._tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self._path = self._tmp.name
        self._tmp.close()  # openpyxl manages file writing
        self._workbook = Workbook(write_only=True)
        self._worksheet = self._workbook.create_sheet()
        self._columns: list[str] = []

    def write_header(self, columns: list[str]) -> None:
        self._columns = columns
        self._worksheet.append(columns)

    def write_row(self, row: dict) -> None:
        values = []
        for col in self._columns:
            value = row.get(col)
            if value is not None and not isinstance(value, str | int | float | bool):
                value = str(value)
            values.append(sanitize_value_for_excel(value))
        self._worksheet.append(values)

    def finish(self) -> str:
        self._workbook.save(self._path)
        return self._path


@dataclass
class RowBuffer:
    file: Any
    columns: list[str]
    seen_keys: set[str]
    row_count: int

    def __iter__(self) -> Iterator[dict]:
        for line in self.file:
            if line.strip():
                yield json.loads(line)


@contextmanager
def _buffer_rows(exported_asset: ExportedAsset, limit: int) -> Iterator[RowBuffer]:
    """Buffer rows to a temp file, discovering columns along the way.

    Yields a RowBuffer that exposes:
    - columns: list of column names (grouped by prefix)
    - seen_keys: set of all seen keys
    - row_count: number of rows
    - __iter__: yields rows as dicts
    """
    with tempfile.NamedTemporaryFile(mode="w+", suffix=".jsonl", delete=True) as jsonl_file:
        row_count, columns, seen_keys = _write_rows_to_jsonl(jsonl_file, exported_asset, limit)
        jsonl_file.seek(0)

        yield RowBuffer(
            file=jsonl_file,
            columns=columns,
            seen_keys=seen_keys,
            row_count=row_count,
        )


def _group_columns_by_prefix(all_keys: list[str]) -> list[str]:
    """Group columns by their top-level prefix.

    Ensures all 'properties.*' columns are grouped together, all 'distinct_ids.*'
    columns are together, etc. Maintains insertion order within each group.
    """
    ordered_fields: OrderedDict[str, list[str]] = OrderedDict()
    for key in all_keys:
        prefix = key.split(".")[0]
        if prefix in ordered_fields:
            ordered_fields[prefix].append(key)
        else:
            ordered_fields[prefix] = [key]

    return [key for group in ordered_fields.values() for key in group]


def add_query_params(url: str, params: dict[str, str]) -> str:
    """
    Uses parse_qsl because parse_qs turns all values into lists but doesn't unbox them when re-encoded
    """
    parsed = urlparse(url)
    query_params = parse_qsl(parsed.query, keep_blank_values=True)

    update_params: list[tuple[str, Any]] = []
    for param, value in query_params:
        if param in params:
            update_params.append((param, params.pop(param)))
        else:
            update_params.append((param, value))

    for key, value in params.items():
        update_params.append((key, value))

    encodedQueryParams = urlencode(update_params, quote_via=quote)
    parsed = parsed._replace(query=encodedQueryParams)
    return urlunparse(parsed)


def _get_breakdown_info(
    item: dict, breakdown_filter: Optional[dict]
) -> tuple[list[str | int | None], list[dict], bool]:
    """Extract breakdown info from query results.

    Supports both breakdown formats from BreakdownFilter (posthog/schema.py)
    """
    breakdown_values: list[str | int | None] = []
    breakdowns: list[dict] = []

    if "breakdown_value" in item:
        breakdown_value = item.get("breakdown_value")
        breakdown_values = breakdown_value if isinstance(breakdown_value, list) else [breakdown_value]

        breakdowns = breakdown_filter.get("breakdowns", []) if breakdown_filter else []
        # Fallback to simplified breakdown field if breakdowns array not present
        if not breakdowns and breakdown_filter and "breakdown" in breakdown_filter:
            breakdowns = [{"property": breakdown_filter.get("breakdown")}]

    has_breakdown_columns = bool(breakdowns)

    return breakdown_values, breakdowns, has_breakdown_columns


def _convert_response_to_csv_data(data: Any, breakdown_filter: Optional[dict] = None) -> Generator[Any, None, None]:
    if isinstance(data.get("results"), list):
        results = data.get("results")
        if len(results) > 0 and (isinstance(results[0], list) or isinstance(results[0], tuple)) and data.get("types"):
            # e.g. {'columns': ['count()'], 'hasMore': False, 'results': [[1775]], 'types': ['UInt64']}
            # or {'columns': ['count()', 'event'], 'hasMore': False, 'results': [[551, '$feature_flag_called'], [265, '$autocapture']], 'types': ['UInt64', 'String']}
            for row in results:
                row_dict = {}
                for idx, x in enumerate(row):
                    if isinstance(x, dict):
                        for key in filter(
                            lambda y: y in RESULT_LIMIT_KEYS and len(x[y]) > RESULT_LIMIT_LENGTH, x.keys()
                        ):
                            total = len(x[key])
                            x[key] = x[key][:RESULT_LIMIT_LENGTH]
                            row_dict[f"{key}.total"] = f"Note: {total} {key} in total"

                    if not data.get("columns"):
                        row_dict[f"column_{idx}"] = x
                    else:
                        row_dict[data["columns"][idx]] = x
                yield row_dict
            return

    if isinstance(data.get("results"), list) or isinstance(data.get("results"), dict):
        results = data.get("results")
    elif isinstance(data.get("result"), list) or isinstance(data.get("result"), dict):
        results = data.get("result")
    else:
        return None

    if isinstance(results, list):
        first_result = next(iter(results), None)

        if not first_result:
            return
        elif len(results) == 1 and isinstance(first_result, dict) and set(first_result.keys()) == {"people", "count"}:
            # persons modal like
            yield from results[0].get("people")
            return
        elif isinstance(first_result, list) or first_result.get("action_id"):
            multiple_items = results if isinstance(first_result, list) else [results]
            # FUNNELS LIKE
            for items in multiple_items:
                yield from (
                    {
                        "name": x.get("custom_name") or x.get("action_id", ""),
                        "breakdown_value": "::".join(x.get("breakdown_value", [])),
                        "action_id": x.get("action_id", ""),
                        "count": x.get("count", ""),
                        "median_conversion_time (seconds)": x.get("median_conversion_time", ""),
                        "average_conversion_time (seconds)": x.get("average_conversion_time", ""),
                    }
                    for x in items
                )
            return
        elif first_result.get("appearances") and first_result.get("person"):
            # RETENTION PERSONS LIKE
            period = data["filters"]["period"] or "Day"
            for item in results:
                line = {"person": item["person"]["name"]}
                for index, data in enumerate(item["appearances"]):
                    line[f"{period} {index}"] = data

                yield line
            return
        elif first_result.get("values") and first_result.get("label"):
            # RETENTION LIKE
            for item in results:
                item_values = item.get("values", [])
                cohort_size = item_values[0]["count"] if item_values else 0

                if item.get("date"):
                    # Dated means we create a grid
                    line = {
                        "cohort": item["date"],
                        "cohort size": cohort_size,
                    }
                    for data in item_values:
                        line[data["label"]] = data["count"]
                else:
                    # Otherwise we just specify "Period" for titles
                    line = {
                        "cohort": item["label"],
                        "cohort size": cohort_size,
                    }
                    for index, data in enumerate(item_values):
                        line[f"Period {index}"] = data["count"]

                yield line
            return
        elif isinstance(first_result.get("data"), list) or (
            first_result.get("data") is None and "aggregated_value" in first_result
        ):
            is_comparison = first_result.get("compare_label")

            # take date labels from current results, when comparing against previous
            # as previous results will be indexed with offset
            date_labels_item = next((x for x in results if x.get("compare_label") == "current"), None)

            # TRENDS LIKE
            for index, item in enumerate(results):
                label = item.get("label", f"Series #{index + 1}")
                compare_label = item.get("compare_label", "")
                action = item.get("action")

                breakdown_values, breakdowns, has_breakdown_columns = _get_breakdown_info(item, breakdown_filter)

                if has_breakdown_columns and isinstance(action, dict) and action.get("name"):
                    series_name = action["name"]
                else:
                    series_name = label

                if compare_label:
                    series_name = f"{series_name} - {compare_label}"

                line = {"series": series_name}

                label_item = date_labels_item if is_comparison else item

                if isinstance(action, dict) and action.get("custom_name"):
                    line["custom name"] = action.get("custom_name")

                for idx, val in enumerate(breakdown_values):
                    prop_name = breakdowns[idx].get("property") if idx < len(breakdowns) else None
                    if not prop_name:
                        continue
                    # Convert list property names to string (e.g., HogQL expressions)
                    if isinstance(prop_name, list):
                        prop_name = ", ".join(str(p) for p in prop_name)
                    formatted_val = str(val) if val is not None else ""
                    if formatted_val == BREAKDOWN_OTHER_STRING_LABEL:
                        formatted_val = BREAKDOWN_OTHER_DISPLAY
                    elif formatted_val == BREAKDOWN_NULL_STRING_LABEL:
                        formatted_val = BREAKDOWN_NULL_DISPLAY
                    line[prop_name] = formatted_val

                if item.get("aggregated_value") is not None:
                    line["Total Sum"] = item.get("aggregated_value")
                elif item.get("data"):
                    labels = label_item.get("labels", []) if label_item else []
                    for index, data in enumerate(item["data"]):
                        if index < len(labels):
                            line[labels[index]] = data

                yield line

            return
    elif results and isinstance(results, dict):
        if "bins" in results:
            for bin_entry in results["bins"]:
                if isinstance(bin_entry, (list, tuple)) and len(bin_entry) >= 2:
                    yield {"bin": bin_entry[0], "value": bin_entry[1]}
            return

    # Pagination object
    yield from results
    return


class UnexpectedEmptyJsonResponse(Exception):
    pass


def get_from_insights_api(exported_asset: ExportedAsset, limit: int, resource: dict) -> Generator[Any, None, None]:
    path: str = resource["path"]
    method: str = resource.get("method", "GET")
    body = resource.get("body", None)
    next_url = None
    access_token = encode_jwt(
        {"id": exported_asset.created_by_id},
        datetime.timedelta(minutes=15),
        PosthogJwtAudience.IMPERSONATED_USER,
    )
    total = 0
    while total < CSV_EXPORT_LIMIT:
        try:
            response = make_api_call(access_token, body, limit, method, next_url, path)
        except HTTPError as e:
            if "Query size exceeded" not in e.response.text:
                raise

            if limit <= CSV_EXPORT_BREAKDOWN_LIMIT_LOW:
                break  # Already tried with the lowest limit, so return what we have

            # If error message contains "Query size exceeded", we try again with a lower limit
            limit = int(limit / 2)
            logger.warning(
                "csv_exporter.query_size_exceeded",
                exc=e,
                exc_info=True,
                response_text=e.response.text,
                limit=limit,
            )
            continue

        # Figure out how to handle funnel polling....
        data = response.json()

        if data is None:
            unexpected_empty_json_response = UnexpectedEmptyJsonResponse("JSON is None when calling API for data")
            logger.error(
                "csv_exporter.json_was_none",
                exc=unexpected_empty_json_response,
                exc_info=True,
                response_text=response.text,
            )

            raise unexpected_empty_json_response

        csv_rows = list(_convert_response_to_csv_data(data))
        total += len(csv_rows)
        yield from csv_rows

        if not data.get("next") or not csv_rows:
            break

        next_url = data.get("next")


def get_from_query(exported_asset: ExportedAsset, limit: int, resource: dict) -> Generator[Any, None, None]:
    query = resource.get("source")
    assert query is not None

    breakdown_filter = query.get("breakdownFilter") if query else None
    total = 0

    # Pagination state - detected from response
    cursor: str | None = None
    offset = 0
    use_cursor = False
    supports_pagination: bool | None = None  # None = not yet detected

    while total < CSV_EXPORT_LIMIT:
        # Build paginated query
        paginated_query = query.copy()

        # Only add pagination parameters after confirming the query supports pagination
        if supports_pagination:
            paginated_query["limit"] = QUERY_PAGE_SIZE
            if cursor is not None:
                paginated_query["after"] = cursor
            elif offset > 0:
                paginated_query["offset"] = offset

        try:
            query_response = process_query_dict(
                team=exported_asset.team,
                query_json=paginated_query,
                limit_context=LimitContext.EXPORT,
                execution_mode=ExecutionMode.CALCULATE_BLOCKING_ALWAYS,
            )
        except QuerySizeExceeded:
            if "breakdownFilter" not in query or limit <= CSV_EXPORT_BREAKDOWN_LIMIT_LOW:
                raise

            # HACKY: Adjust the breakdown_limit in the query
            limit = int(limit / 2)
            query["breakdownFilter"]["breakdown_limit"] = limit
            continue

        if isinstance(query_response, BaseModel):
            response_dict = query_response.model_dump(by_alias=True)
        else:
            response_dict = query_response

        rows = list(_convert_response_to_csv_data(response_dict, breakdown_filter=breakdown_filter))
        rows = rows[: CSV_EXPORT_LIMIT - total]
        total += len(rows)
        yield from rows

        if total >= CSV_EXPORT_LIMIT or len(rows) == 0:
            break

        # Detect pagination support from response
        next_cursor = response_dict.get("nextCursor") or response_dict.get("next_cursor")
        has_more = response_dict.get("hasMore", False)

        if next_cursor:
            # Priority 1: Cursor pagination
            cursor = next_cursor
            use_cursor = True
            supports_pagination = True
        elif has_more and not use_cursor:
            # Priority 2: Offset pagination
            offset += QUERY_PAGE_SIZE
            supports_pagination = True
        else:
            # No pagination indicators - single query only
            break


def _iter_rows(exported_asset: ExportedAsset, limit: int) -> Generator[Any, None, None]:
    resource = exported_asset.export_context or {}

    if resource.get("source"):
        yield from get_from_query(exported_asset, limit, resource)
    else:
        # Legacy path for PersonsNode exports (uses API path instead of HogQL source).
        # PersonsNode was migrated to ActorsQuery in migration 0459, so this path
        # should rarely be hit in practice.
        yield from get_from_insights_api(exported_asset, CSV_EXPORT_BREAKDOWN_LIMIT_INITIAL, resource)


def _write_rows_to_jsonl(jsonl_file: Any, exported_asset: ExportedAsset, limit: int) -> tuple[int, list[str], set[str]]:
    """Write flattened rows to a JSON lines file, discovering columns as we go.

    Returns:
        Tuple of (row_count, all_keys, seen_keys)
    """
    renderer = CSVRenderer()
    all_keys: list[str] = []
    seen_keys: set[str] = set()
    row_count = 0

    for row in _iter_rows(exported_asset, limit):
        flat_row = dict(renderer.flatten_item(row))

        for key in flat_row.keys():
            if key not in seen_keys:
                seen_keys.add(key)
                all_keys.append(key)

        json.dump(flat_row, jsonl_file, default=str)
        jsonl_file.write("\n")
        row_count += 1

    grouped_keys = _group_columns_by_prefix(all_keys)

    return row_count, grouped_keys, seen_keys


def _determine_columns(user_columns: list[str], all_keys: list[str], seen_keys: set[str]) -> list[str]:
    """Determine the final column order based on user preferences and discovered keys."""
    if not user_columns:
        return all_keys

    columns = []
    for col in user_columns:
        if col in seen_keys:
            columns.append(col)
        else:
            # Check if there are nested keys that start with this prefix
            nested_keys = [key for key in all_keys if key.startswith(col + ".")]
            if nested_keys:
                columns.extend(nested_keys)
            else:
                # Include the column even if it doesn't exist in data (will be empty)
                columns.append(col)
    return columns


def _export_tabular(exported_asset: ExportedAsset, limit: int, writer: TabularWriter) -> None:
    """Export data using the provided writer."""
    user_columns = (exported_asset.export_context or {}).get("columns", [])

    with _buffer_rows(exported_asset, limit) as buffer:
        if buffer.row_count == 0:
            columns = user_columns if user_columns else ["error"]
            writer.write_header(columns)
            if user_columns:
                writer.write_row({})  # empty row
            else:
                writer.write_row({"error": "No data available or unable to format for export."})
        else:
            columns = _determine_columns(user_columns, buffer.columns, buffer.seen_keys)
            writer.write_header(columns)
            for row in buffer:
                writer.write_row(row)

    path = writer.finish()
    try:
        save_content_from_file(exported_asset, path)
    finally:
        os.unlink(path)


def get_limit_param_key(path: str) -> str:
    query = QueryDict(path)
    breakdown = query.get("breakdown", None)
    return "breakdown_limit" if breakdown is not None else "limit"


def make_api_call(
    access_token: str,
    body: Any,
    limit: int,
    method: str,
    next_url: Optional[str],
    path: str,
) -> requests.models.Response:
    request_url: str = absolute_uri(next_url or path)
    url = add_query_params(
        request_url,
        {get_limit_param_key(request_url): str(limit), "is_csv_export": "1"},
    )
    response = requests.request(
        method=method.lower(),
        url=url,
        json=body,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=60,
    )
    response.raise_for_status()
    return response


def export_tabular(exported_asset: ExportedAsset, limit: Optional[int] = None) -> None:
    if not limit:
        limit = CSV_EXPORT_BREAKDOWN_LIMIT_INITIAL

    try:
        with EXPORT_TIMER.labels(type=exported_asset.export_format).time():
            if exported_asset.export_format == ExportedAsset.ExportFormat.CSV:
                _export_tabular(exported_asset, limit, CsvWriter())
            elif exported_asset.export_format == ExportedAsset.ExportFormat.XLSX:
                _export_tabular(exported_asset, limit, ExcelWriter())
            else:
                raise NotImplementedError(f"Export to format {exported_asset.export_format} is not supported")
    except Exception as e:
        if exported_asset:
            team_id = str(exported_asset.team.id)
        else:
            team_id = "unknown"

        capture_exception(e, additional_properties={"task": "csv_export", "team_id": team_id})
        logger.error("csv_exporter.failed", exception=e, exc_info=True)
        raise
