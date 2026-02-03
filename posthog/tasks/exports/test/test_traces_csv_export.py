"""Tests for TracesQuery CSV/XLSX export functionality."""

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from posthog.test.base import APIBaseTest, ClickhouseTestMixin, _create_event, _create_person, flush_persons_and_events
from unittest.mock import patch

from django.test import override_settings

from openpyxl import load_workbook

from posthog.models import ExportedAsset, PropertyDefinition
from posthog.models.property_definition import PropertyType
from posthog.storage.object_storage import ObjectStorageError
from posthog.tasks.exports import csv_exporter

TEST_PREFIX = "Test-Exports"


def _create_ai_generation_event(
    *,
    team,
    distinct_id: str,
    trace_id: str,
    input_messages: list[dict[str, str]] | None = None,
    output_messages: list[dict[str, str]] | None = None,
    timestamp: datetime | None = None,
    properties: dict[str, Any] | None = None,
):
    """Create an $ai_generation event."""
    if input_messages is None:
        input_messages = [{"role": "user", "content": "Hello"}]
    if output_messages is None:
        output_messages = [{"role": "assistant", "content": "Hi there!"}]

    props = {
        "$ai_trace_id": trace_id,
        # Set parent_id to trace_id so the event is included in the trace's events array
        "$ai_parent_id": trace_id,
        "$ai_latency": 1.5,
        "$ai_input": input_messages,
        "$ai_output_choices": output_messages,
        "$ai_input_tokens": 10,
        "$ai_output_tokens": 15,
        "$ai_total_cost_usd": 0.001,
    }
    if properties:
        props.update(properties)

    _create_event(
        event="$ai_generation",
        distinct_id=distinct_id,
        properties=props,
        team=team,
        timestamp=timestamp or datetime.now(UTC),
    )


def _create_ai_trace_event(
    *,
    team,
    distinct_id: str,
    trace_id: str,
    trace_name: str | None = None,
    input_state: dict[str, Any] | None = None,
    output_state: dict[str, Any] | None = None,
    timestamp: datetime | None = None,
):
    """Create an $ai_trace event with input/output state."""
    props: dict[str, Any] = {
        "$ai_trace_id": trace_id,
        "$ai_span_name": trace_name,
    }
    # Store input_state and output_state as dicts - ClickHouse will serialize them to JSON
    if input_state is not None:
        props["$ai_input_state"] = input_state
    if output_state is not None:
        props["$ai_output_state"] = output_state

    _create_event(
        event="$ai_trace",
        distinct_id=distinct_id,
        properties=props,
        team=team,
        timestamp=timestamp or datetime.now(UTC),
    )


@override_settings(SITE_URL="http://testserver")
class TestTracesQueryCSVExport(ClickhouseTestMixin, APIBaseTest):
    """Test TracesQuery CSV/XLSX export with inputState and outputState."""

    def setUp(self):
        super().setUp()
        self._create_properties()

    def _create_properties(self):
        """Create property definitions for numeric AI properties."""
        numeric_props = {
            "$ai_latency",
            "$ai_input_tokens",
            "$ai_output_tokens",
            "$ai_total_cost_usd",
        }
        models_to_create = []
        for prop in numeric_props:
            prop_model = PropertyDefinition(
                team=self.team,
                name=prop,
                type=PropertyDefinition.Type.EVENT,
                property_type=PropertyType.Numeric,
            )
            models_to_create.append(prop_model)
        PropertyDefinition.objects.bulk_create(models_to_create)

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write_from_file")
    def test_traces_export_csv_includes_input_output_state(
        self, mocked_object_storage_write_from_file: Any, mocked_uuidt: Any
    ) -> None:
        """Test that CSV export of TracesQuery includes inputState and outputState content."""
        _create_person(distinct_ids=["user1"], team=self.team)

        trace_id = str(uuid.uuid4())

        # Create a trace with input/output state
        _create_ai_trace_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            trace_name="test_conversation",
            input_state={"messages": [{"role": "user", "content": "What is the weather?"}]},
            output_state={
                "messages": [
                    {"role": "user", "content": "What is the weather?"},
                    {"role": "assistant", "content": "The weather is sunny today."},
                ]
            },
            timestamp=datetime.now(UTC),
        )

        # Create a generation event (needed to have a trace)
        _create_ai_generation_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            timestamp=datetime.now(UTC),
        )

        flush_persons_and_events()

        # Create export asset with columns including inputState and outputState
        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "columns": ["id", "traceName", "inputState", "outputState"],
                "source": {
                    "kind": "TracesQuery",
                    "dateRange": {
                        "date_from": "-1d",
                        "date_to": None,
                    },
                },
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "test-guid"
        mocked_object_storage_write_from_file.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            # When S3 fails, content is stored directly in the asset
            content = exported_asset.content.decode("utf-8")
            assert content is not None, "Export content should not be None"

            lines = content.strip().split("\r\n")
            assert len(lines) >= 2, f"Should have at least header and one data row. Content: {content}"

            # Check that inputState and outputState data is in the export
            # The content should include the actual message content, not just empty columns
            assert "What is the weather?" in content, f"Input message not found in export. Content: {content}"
            assert "The weather is sunny today" in content, f"Output message not found in export. Content: {content}"

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write_from_file")
    def test_traces_export_xlsx_includes_input_output_state(
        self, mocked_object_storage_write_from_file: Any, mocked_uuidt: Any
    ) -> None:
        """Test that XLSX export of TracesQuery includes inputState and outputState content."""
        _create_person(distinct_ids=["user1"], team=self.team)

        trace_id = str(uuid.uuid4())

        # Create a trace with input/output state
        _create_ai_trace_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            trace_name="test_conversation",
            input_state={"messages": [{"role": "user", "content": "Tell me a joke"}]},
            output_state={
                "messages": [
                    {"role": "user", "content": "Tell me a joke"},
                    {"role": "assistant", "content": "Why did the chicken cross the road?"},
                ]
            },
            timestamp=datetime.now(UTC),
        )

        # Create a generation event
        _create_ai_generation_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            timestamp=datetime.now(UTC),
        )

        flush_persons_and_events()

        # Create export asset with columns including inputState and outputState
        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.XLSX,
            export_context={
                "columns": ["id", "traceName", "inputState", "outputState"],
                "source": {
                    "kind": "TracesQuery",
                    "dateRange": {
                        "date_from": "-1d",
                        "date_to": None,
                    },
                },
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "test-guid"
        mocked_object_storage_write_from_file.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            # When S3 fails, content is stored directly in the asset
            content = exported_asset.content
            assert content is not None, "Export content should not be None"

            # Parse the Excel file
            workbook = load_workbook(filename=BytesIO(content))
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))

            assert len(rows) >= 2, f"Should have at least header and one data row. Rows: {rows}"

            # Convert all cell values to string for checking
            all_content = " ".join(str(cell) for row in rows for cell in row if cell is not None)

            # Check that inputState and outputState data is in the export
            assert "Tell me a joke" in all_content, f"Input message not found in export. Content: {all_content}"
            assert "Why did the chicken cross the road?" in all_content, (
                f"Output message not found in export. Content: {all_content}"
            )

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write_from_file")
    def test_traces_export_without_trace_event_has_null_states(
        self, mocked_object_storage_write_from_file: Any, mocked_uuidt: Any
    ) -> None:
        """Test that export handles traces without $ai_trace events (inputState/outputState will be null)."""
        _create_person(distinct_ids=["user1"], team=self.team)

        trace_id = str(uuid.uuid4())

        # Create only a generation event (no trace event)
        _create_ai_generation_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            input_messages=[{"role": "user", "content": "Test input"}],
            output_messages=[{"role": "assistant", "content": "Test output"}],
            timestamp=datetime.now(UTC),
        )

        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "columns": ["id", "inputState", "outputState"],
                "source": {
                    "kind": "TracesQuery",
                    "dateRange": {
                        "date_from": "-1d",
                        "date_to": None,
                    },
                },
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "test-guid"
        mocked_object_storage_write_from_file.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            # When S3 fails, content is stored directly in the asset
            content = exported_asset.content.decode("utf-8")
            assert content is not None, "Export content should not be None"

            # Without an $ai_trace event, inputState and outputState will be null
            # The export should still succeed with empty columns for these fields
            lines = content.strip().split("\r\n")
            assert len(lines) >= 2, f"Should have at least header and one data row. Content: {content}"

            # Verify the header includes the requested columns
            header = lines[0]
            assert "inputState" in header, f"inputState column not found in header. Content: {content}"
            assert "outputState" in header, f"outputState column not found in header. Content: {content}"

            # The trace ID should still be present
            assert trace_id in content, f"Trace ID not found in export. Content: {content}"

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write_from_file")
    def test_traces_export_all_columns_discovers_nested_state(
        self, mocked_object_storage_write_from_file: Any, mocked_uuidt: Any
    ) -> None:
        """Test that exporting without specific columns discovers nested inputState/outputState fields."""
        _create_person(distinct_ids=["user1"], team=self.team)

        trace_id = str(uuid.uuid4())

        _create_ai_trace_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            trace_name="discovery_test",
            input_state={"messages": [{"role": "user", "content": "Discovery test input"}]},
            output_state={"messages": [{"role": "assistant", "content": "Discovery test output"}]},
            timestamp=datetime.now(UTC),
        )

        _create_ai_generation_event(
            team=self.team,
            distinct_id="user1",
            trace_id=trace_id,
            timestamp=datetime.now(UTC),
        )

        flush_persons_and_events()

        # Export without specifying columns - should discover all fields
        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "TracesQuery",
                    "dateRange": {
                        "date_from": "-1d",
                        "date_to": None,
                    },
                },
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "test-guid"
        mocked_object_storage_write_from_file.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            # When S3 fails, content is stored directly in the asset
            content = exported_asset.content.decode("utf-8")
            assert content is not None, "Export content should not be None"

            # The export should include the nested message content
            assert "Discovery test input" in content, f"Input message not found in export. Content: {content}"
            assert "Discovery test output" in content, f"Output message not found in export. Content: {content}"
