from datetime import datetime
from io import BytesIO
from typing import Any, Optional

import pytest
from freezegun import freeze_time
from posthog.test.base import APIBaseTest, _create_event, _create_person, flush_persons_and_events
from unittest import mock
from unittest.mock import ANY, MagicMock, Mock, patch

from django.test import override_settings
from django.utils.timezone import now

from boto3 import resource
from botocore.client import Config
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from requests.exceptions import HTTPError

from posthog.hogql.constants import CSV_EXPORT_BREAKDOWN_LIMIT_INITIAL

from posthog.models import ExportedAsset
from posthog.models.utils import UUIDT
from posthog.settings import (
    OBJECT_STORAGE_ACCESS_KEY_ID,
    OBJECT_STORAGE_BUCKET,
    OBJECT_STORAGE_ENDPOINT,
    OBJECT_STORAGE_SECRET_ACCESS_KEY,
)
from posthog.storage import object_storage
from posthog.storage.object_storage import ObjectStorageError
from posthog.tasks.exports import csv_exporter
from posthog.tasks.exports.csv_exporter import (
    UnexpectedEmptyJsonResponse,
    _convert_response_to_csv_data,
    add_query_params,
)
from posthog.test.test_journeys import journeys_for
from posthog.utils import absolute_uri

TEST_PREFIX = "Test-Exports"

# see GitHub issue #11204
regression_11204 = "api/projects/6642/insights/trend/?events=%5B%7B%22id%22%3A%22product%20viewed%22%2C%22name%22%3A%22product%20viewed%22%2C%22type%22%3A%22events%22%2C%22order%22%3A0%7D%5D&actions=%5B%5D&display=ActionsTable&insight=TRENDS&interval=day&breakdown=productName&new_entity=%5B%5D&properties=%5B%5D&step_limit=5&funnel_filter=%7B%7D&breakdown_type=event&exclude_events=%5B%5D&path_groupings=%5B%5D&include_event_types=%5B%22%24pageview%22%5D&filter_test_accounts=false&local_path_cleaning_filters=%5B%5D&date_from=-14d&offset=50"


@override_settings(SITE_URL="http://testserver")
class TestCSVExporter(APIBaseTest):
    @pytest.fixture(autouse=True)
    def patched_request(self):
        with patch("posthog.tasks.exports.csv_exporter.requests.request") as patched_request:
            mock_response = Mock()
            mock_response.status_code = 200
            # API responses copied from https://github.com/PostHog/posthog/runs/7221634689?check_suite_focus=true
            mock_response.json.side_effect = [
                {
                    "next": "http://testserver/api/projects/169/events?orderBy=%5B%22-timestamp%22%5D&properties=%5B%7B%22key%22%3A%22%24browser%22%2C%22value%22%3A%5B%22Safari%22%5D%2C%22operator%22%3A%22exact%22%2C%22type%22%3A%22event%22%7D%5D&after=2022-07-06T19%3A27%3A43.206326&limit=1&before=2022-07-06T19%3A37%3A43.095295%2B00%3A00",
                    "results": [
                        {
                            "id": "e9ca132e-400f-4854-a83c-16c151b2f145",
                            "distinct_id": "2",
                            "properties": {"$browser": "Safari"},
                            "event": "event_name",
                            "timestamp": "2022-07-06T19:37:43.095295+00:00",
                            "person": None,
                            "elements": [],
                            "elements_chain": "",
                        }
                    ],
                },
                {
                    "next": "http://testserver/api/projects/169/events?orderBy=%5B%22-timestamp%22%5D&properties=%5B%7B%22key%22%3A%22%24browser%22%2C%22value%22%3A%5B%22Safari%22%5D%2C%22operator%22%3A%22exact%22%2C%22type%22%3A%22event%22%7D%5D&after=2022-07-06T19%3A27%3A43.206326&limit=1&before=2022-07-06T19%3A37%3A43.095279%2B00%3A00",
                    "results": [
                        {
                            "id": "1624228e-a4f1-48cd-aabc-6baa3ddb22e4",
                            "distinct_id": "2",
                            "properties": {"$browser": "Safari"},
                            "event": "event_name",
                            "timestamp": "2022-07-06T19:37:43.095279+00:00",
                            "person": None,
                            "elements": [],
                            "elements_chain": "",
                        }
                    ],
                },
                {
                    "next": None,
                    "results": [
                        {
                            "id": "66d45914-bdf5-4980-a54a-7dc699bdcce9",
                            "distinct_id": "2",
                            "properties": {"$browser": "Safari"},
                            "event": "event_name",
                            "timestamp": "2022-07-06T19:37:43.095262+00:00",
                            "person": None,
                            "elements": [],
                            "elements_chain": "",
                        }
                    ],
                },
            ]
            patched_request.return_value = mock_response
            yield patched_request

    def _create_asset(self, extra_context: Optional[dict] = None) -> ExportedAsset:
        if extra_context is None:
            extra_context = {}

        asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={"path": "/api/literally/anything", **extra_context},
        )
        asset.save()
        return asset

    def _split_to_dict(self, url: str) -> dict[str, Any]:
        first_split_parts = url.split("?")
        assert len(first_split_parts) == 2
        return {bits[0]: bits[1] for bits in [param.split("=") for param in first_split_parts[1].split("&")]}

    def teardown_method(self, method):
        s3 = resource(
            "s3",
            endpoint_url=OBJECT_STORAGE_ENDPOINT,
            aws_access_key_id=OBJECT_STORAGE_ACCESS_KEY_ID,
            aws_secret_access_key=OBJECT_STORAGE_SECRET_ACCESS_KEY,
            config=Config(signature_version="s3v4"),
            region_name="us-east-1",
        )
        bucket = s3.Bucket(OBJECT_STORAGE_BUCKET)
        bucket.objects.filter(Prefix=TEST_PREFIX).delete()

    def test_csv_exporter_writes_to_asset_when_object_storage_is_disabled(self) -> None:
        exported_asset = self._create_asset()
        with self.settings(OBJECT_STORAGE_ENABLED=False):
            csv_exporter.export_tabular(exported_asset)

            assert (
                exported_asset.content
                == b"id,distinct_id,properties.$browser,event,timestamp,person,elements_chain\r\ne9ca132e-400f-4854-a83c-16c151b2f145,2,Safari,event_name,2022-07-06T19:37:43.095295+00:00,,\r\n1624228e-a4f1-48cd-aabc-6baa3ddb22e4,2,Safari,event_name,2022-07-06T19:37:43.095279+00:00,,\r\n66d45914-bdf5-4980-a54a-7dc699bdcce9,2,Safari,event_name,2022-07-06T19:37:43.095262+00:00,,\r\n"
            )
            assert exported_asset.content_location is None

    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_writes_to_object_storage_when_object_storage_is_enabled(self, mocked_uuidt) -> None:
        exported_asset = self._create_asset()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert (
                exported_asset.content_location
                == f"{TEST_PREFIX}/csv/team-{self.team.id}/task-{exported_asset.id}/a-guid"
            )

            content = object_storage.read(exported_asset.content_location)
            assert (
                content
                == "id,distinct_id,properties.$browser,event,timestamp,person,elements_chain\r\ne9ca132e-400f-4854-a83c-16c151b2f145,2,Safari,event_name,2022-07-06T19:37:43.095295+00:00,,\r\n1624228e-a4f1-48cd-aabc-6baa3ddb22e4,2,Safari,event_name,2022-07-06T19:37:43.095279+00:00,,\r\n66d45914-bdf5-4980-a54a-7dc699bdcce9,2,Safari,event_name,2022-07-06T19:37:43.095262+00:00,,\r\n"
            )

            assert exported_asset.content is None

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_writes_to_asset_when_object_storage_write_fails(
        self, mocked_object_storage_write, mocked_uuidt
    ) -> None:
        exported_asset = self._create_asset()
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert (
                exported_asset.content
                == b"id,distinct_id,properties.$browser,event,timestamp,person,elements_chain\r\ne9ca132e-400f-4854-a83c-16c151b2f145,2,Safari,event_name,2022-07-06T19:37:43.095295+00:00,,\r\n1624228e-a4f1-48cd-aabc-6baa3ddb22e4,2,Safari,event_name,2022-07-06T19:37:43.095279+00:00,,\r\n66d45914-bdf5-4980-a54a-7dc699bdcce9,2,Safari,event_name,2022-07-06T19:37:43.095262+00:00,,\r\n"
            )

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_does_not_filter_columns_on_empty_param(
        self, mocked_object_storage_write, mocked_uuidt
    ) -> None:
        exported_asset = self._create_asset({"columns": []})
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert (
                exported_asset.content
                == b"id,distinct_id,properties.$browser,event,timestamp,person,elements_chain\r\ne9ca132e-400f-4854-a83c-16c151b2f145,2,Safari,event_name,2022-07-06T19:37:43.095295+00:00,,\r\n1624228e-a4f1-48cd-aabc-6baa3ddb22e4,2,Safari,event_name,2022-07-06T19:37:43.095279+00:00,,\r\n66d45914-bdf5-4980-a54a-7dc699bdcce9,2,Safari,event_name,2022-07-06T19:37:43.095262+00:00,,\r\n"
            )

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_does_filter_columns(self, mocked_object_storage_write, mocked_uuidt) -> None:
        # NB these columns are not in the "natural" order
        exported_asset = self._create_asset({"columns": ["distinct_id", "properties.$browser", "event"]})
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert (
                exported_asset.content
                == b"distinct_id,properties.$browser,event\r\n2,Safari,event_name\r\n2,Safari,event_name\r\n2,Safari,event_name\r\n"
            )

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_includes_whole_dict(self, mocked_object_storage_write, mocked_uuidt) -> None:
        exported_asset = self._create_asset({"columns": ["distinct_id", "properties"]})
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert exported_asset.content == b"distinct_id,properties.$browser\r\n2,Safari\r\n2,Safari\r\n2,Safari\r\n"

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_includes_whole_dict_alternative_order(
        self, mocked_object_storage_write, mocked_uuidt
    ) -> None:
        exported_asset = self._create_asset({"columns": ["properties", "distinct_id"]})
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert exported_asset.content == b"properties.$browser,distinct_id\r\nSafari,2\r\nSafari,2\r\nSafari,2\r\n"

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_does_filter_columns_and_can_handle_unexpected_columns(
        self, mocked_object_storage_write, mocked_uuidt
    ) -> None:
        # NB these columns are not in the "natural" order
        exported_asset = self._create_asset({"columns": ["distinct_id", "properties.$browser", "event", "tomato"]})
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.content_location is None

            assert (
                exported_asset.content
                == b"distinct_id,properties.$browser,event,tomato\r\n2,Safari,event_name,\r\n2,Safari,event_name,\r\n2,Safari,event_name,\r\n"
            )

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    def test_csv_exporter_excel(self, mocked_object_storage_write: Any, mocked_uuidt: Any) -> None:
        exported_asset = self._create_asset({"columns": ["distinct_id", "properties.$browser", "event", "tomato"]})
        exported_asset.export_format = ExportedAsset.ExportFormat.XLSX
        mocked_uuidt.return_value = "a-guid"
        mocked_object_storage_write.side_effect = ObjectStorageError("mock write failed")

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert exported_asset.filename == "export.xlsx"
            assert exported_asset.content_location is None

            wb = load_workbook(filename=BytesIO(exported_asset.content))
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            assert data == [
                ("distinct_id", "properties.$browser", "event", "tomato"),
                ("2", "Safari", "event_name", None),
                ("2", "Safari", "event_name", None),
                ("2", "Safari", "event_name", None),
            ]

    @patch("posthog.models.exported_asset.UUIDT")
    @patch("posthog.models.exported_asset.object_storage.write")
    @patch("requests.request")
    def test_csv_exporter_limits_breakdown_insights_correctly(
        self, mocked_request, mocked_object_storage_write, mocked_uuidt
    ) -> None:
        path = "api/projects/1/insights/trend/?insight=TRENDS&breakdown=email&date_from=-7d"
        exported_asset = self._create_asset({"path": path})
        mock_response = Mock()
        mock_response.status_code = 200
        mocked_request.return_value = mock_response

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

        mocked_request.assert_called_with(
            method="get",
            url="http://testserver/" + path + f"&breakdown_limit={CSV_EXPORT_BREAKDOWN_LIMIT_INITIAL}&is_csv_export=1",
            timeout=60,
            json=None,
            headers=ANY,
        )

    @patch("posthog.tasks.exports.csv_exporter.logger")
    def test_failing_export_api_is_reported(self, _mock_logger: MagicMock) -> None:
        with patch("posthog.tasks.exports.csv_exporter.requests.request") as patched_request:
            exported_asset = self._create_asset()
            mock_response = MagicMock()
            mock_response.status_code = 403
            mock_response.raise_for_status.side_effect = Exception("HTTP 403 Forbidden")
            mock_response.ok = False
            patched_request.return_value = mock_response

            with pytest.raises(Exception, match="HTTP 403 Forbidden"):
                csv_exporter.export_tabular(exported_asset)

    @patch("posthog.tasks.exports.csv_exporter.logger")
    def test_failing_export_api_is_reported_query_size_exceeded(self, _mock_logger: MagicMock) -> None:
        with patch("posthog.tasks.exports.csv_exporter.make_api_call") as patched_make_api_call:
            exported_asset = self._create_asset()
            mock_error = HTTPError("Query size exceeded")  # type: ignore[call-arg]
            mock_error.response = Mock()
            mock_error.response.text = "Query size exceeded"
            patched_make_api_call.side_effect = mock_error

            csv_exporter.export_tabular(exported_asset)

            assert patched_make_api_call.call_count == 4
            patched_make_api_call.assert_called_with(mock.ANY, mock.ANY, 64, mock.ANY, mock.ANY, mock.ANY)

    def test_limiting_query_as_expected(self) -> None:
        with self.settings(SITE_URL="https://app.posthog.com"):
            modified_url = add_query_params(absolute_uri(regression_11204), {"limit": "3500"})
            actual_bits = self._split_to_dict(modified_url)
            expected_bits = {
                **self._split_to_dict(regression_11204),
                **{"limit": "3500"},
            }
            assert expected_bits == actual_bits

    def test_limiting_existing_limit_query_as_expected(self) -> None:
        with self.settings(SITE_URL="https://app.posthog.com"):
            url_with_existing_limit = regression_11204 + "&limit=100000"
            modified_url = add_query_params(absolute_uri(url_with_existing_limit), {"limit": "3500"})
            actual_bits = self._split_to_dict(modified_url)
            expected_bits = {
                **self._split_to_dict(regression_11204),
                **{"limit": "3500"},
            }
            assert expected_bits == actual_bits

    @patch("posthog.tasks.exports.csv_exporter.make_api_call")
    def test_raises_expected_error_when_json_is_none(self, patched_api_call) -> None:
        mock_response = Mock()
        mock_response.json.return_value = None
        mock_response.status_code = 200
        mock_response.text = "i am the text"
        patched_api_call.return_value = mock_response

        with pytest.raises(UnexpectedEmptyJsonResponse, match="JSON is None when calling API for data"):
            csv_exporter.export_tabular(self._create_asset())

    @patch("posthog.hogql.constants.CSV_EXPORT_LIMIT", 10)
    @patch("posthog.hogql.constants.DEFAULT_RETURNED_ROWS", 5)
    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_hogql_query(self, mocked_uuidt: Any, DEFAULT_RETURNED_ROWS=5, CSV_EXPORT_LIMIT=10) -> None:
        random_uuid = f"RANDOM_TEST_ID::{UUIDT()}"
        for i in range(15):
            _create_event(
                event="$pageview",
                distinct_id=random_uuid,
                team=self.team,
                timestamp=now() - relativedelta(hours=1),
                properties={"prop": i},
            )
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "HogQLQuery",
                    "query": f"select event from events where distinct_id = '{random_uuid}'",
                }
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)

            assert (
                exported_asset.content_location
                == f"{TEST_PREFIX}/csv/team-{self.team.id}/task-{exported_asset.id}/a-guid"
            )

            content = object_storage.read(exported_asset.content_location)
            assert (
                content
                == "event\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n$pageview\r\n"
            )

            assert exported_asset.content is None

    @patch("posthog.hogql.constants.CSV_EXPORT_LIMIT", 10)
    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_events_query(self, mocked_uuidt: Any, CSV_EXPORT_LIMIT=10) -> None:
        random_uuid = f"RANDOM_TEST_ID::{UUIDT()}"
        for i in range(15):
            _create_event(
                event="$pageview",
                distinct_id=random_uuid,
                team=self.team,
                timestamp=now() - relativedelta(hours=1),
                properties={"prop": i},
            )
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "EventsQuery",
                    "select": ["event", "*"],
                    "where": [f"distinct_id = '{random_uuid}'"],
                }
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)
            lines = (content or "").split("\r\n")
            self.assertEqual(len(lines), 12)
            self.assertEqual(
                lines[0],
                "event,*.uuid,*.event,*.properties.prop,*.timestamp,*.team_id,*.distinct_id,*.elements_chain,*.created_at",
            )
            self.assertEqual(lines[11], "")
            first_row = lines[1].split(",")
            self.assertEqual(first_row[0], "$pageview")
            self.assertEqual(first_row[2], "$pageview")
            self.assertEqual(first_row[5], str(self.team.pk))

    @patch("posthog.hogql.constants.CSV_EXPORT_LIMIT", 10)
    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_events_query_with_columns(self, mocked_uuidt: Any, CSV_EXPORT_LIMIT: int = 10) -> None:
        random_uuid = f"RANDOM_TEST_ID::{UUIDT()}"
        for i in range(15):
            _create_event(
                event="$pageview",
                distinct_id=random_uuid,
                team=self.team,
                timestamp=now() - relativedelta(hours=1),
                properties={"prop": i},
            )
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "columns": ["*"],
                "source": {
                    "kind": "EventsQuery",
                    "select": ["event", "*"],
                    "where": [f"distinct_id = '{random_uuid}'"],
                },
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)
            lines = (content or "").split("\r\n")
            self.assertEqual(len(lines), 12)
            self.assertEqual(
                lines[0],
                "*.uuid,*.event,*.properties.prop,*.timestamp,*.team_id,*.distinct_id,*.elements_chain,*.created_at",
            )
            self.assertEqual(lines[11], "")
            first_row = lines[1].split(",")
            self.assertEqual(first_row[1], "$pageview")
            self.assertEqual(first_row[4], str(self.team.pk))

    @patch("posthog.hogql.constants.MAX_SELECT_RETURNED_ROWS", 10)
    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_funnels_query(self, mocked_uuidt: Any, MAX_SELECT_RETURNED_ROWS: int = 10) -> None:
        _create_person(
            distinct_ids=[f"user_1"],
            team=self.team,
        )

        events_by_person = {
            "user_1": [
                {
                    "event": "$pageview",
                    "timestamp": datetime(2024, 3, 22, 13, 46),
                    "properties": {"utm_medium": "test''123"},
                },
                {
                    "event": "$pageview",
                    "timestamp": datetime(2024, 3, 22, 13, 47),
                    "properties": {"utm_medium": "test''123"},
                },
            ],
        }
        journeys_for(events_by_person, self.team)
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "FunnelsQuery",
                    "series": [
                        {"kind": "EventsNode", "name": "$pageview", "event": "$pageview"},
                        {"kind": "EventsNode", "name": "$pageview", "event": "$pageview"},
                    ],
                    "interval": "day",
                    "dateRange": {"date_to": "2024-03-22", "date_from": "2024-03-22"},
                    "funnelsFilter": {"funnelVizType": "steps"},
                    "breakdownFilter": {"breakdown": "utm_medium", "breakdown_type": "event"},
                }
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)
            lines = (content or "").strip().split("\r\n")
            self.assertEqual(
                lines,
                [
                    "name,breakdown_value,action_id,count,median_conversion_time (seconds),average_conversion_time (seconds)",
                    "$pageview,test'123,$pageview,1,,",
                    "$pageview,test'123,$pageview,1,60.0,60.0",
                ],
            )

    def test_funnel_time_to_convert(self) -> None:
        bins = [
            [1, 1],
            [2, 3],
            [3, 5],
            [4, 17],
            [5, 29],
            [6, 44],
            [7, 38],
            [8, 24],
            [9, 10],
            [10, 7],
            [11, 3],
            [12, 1],
            [13, 1],
            [14, 1],
            [15, 0],
            [16, 0],
            [17, 0],
            [18, 0],
            [19, 0],
            [20, 0],
            [21, 1],
            [22, 0],
            [23, 1],
            [24, 0],
            [25, 0],
            [26, 0],
        ]
        data = {
            "results": {
                "average_conversion_time": 1.45,
                "bins": bins,
            }
        }
        csv_list = list(_convert_response_to_csv_data(data))
        assert len(bins) == len(csv_list)
        for bin, csv in zip(bins, csv_list):
            assert bin[0] == csv["bin"]
            assert bin[1] == csv["value"]

    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_empty_result(self, mocked_uuidt: Any) -> None:
        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "FunnelsQuery",
                    "series": [
                        {"kind": "EventsNode", "name": "$pageview", "event": "$pageview"},
                        {"kind": "EventsNode", "name": "$pageview", "event": "$pageview"},
                    ],
                    "interval": "day",
                    "dateRange": {"date_to": "2024-03-22", "date_from": "2024-03-22"},
                    "funnelsFilter": {"funnelVizType": "steps"},
                    "breakdownFilter": {"breakdown": "utm_medium", "breakdown_type": "event"},
                }
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with patch("posthog.tasks.exports.csv_exporter.get_from_hogql_query") as mocked_get_from_hogql_query:
            mocked_get_from_hogql_query.return_value = iter([])

            with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
                csv_exporter.export_tabular(exported_asset)
                content = object_storage.read(exported_asset.content_location)
                lines = (content or "").split("\r\n")
                self.assertEqual(lines[0], "error")
                self.assertEqual(lines[1], "No data available or unable to format for export.")

    @patch("posthog.hogql.constants.MAX_SELECT_RETURNED_ROWS", 10)
    @patch("posthog.models.exported_asset.UUIDT")
    def test_csv_exporter_trends_query_with_none_action(
        self, mocked_uuidt: Any, MAX_SELECT_RETURNED_ROWS: int = 10
    ) -> None:
        _create_person(
            distinct_ids=[f"user_1"],
            team=self.team,
        )

        events_by_person = {
            "user_1": [
                {
                    "event": "$pageview",
                    "timestamp": datetime(2024, 3, 22, 13, 46),
                },
                {
                    "event": "$pageview",
                    "timestamp": datetime(2024, 3, 22, 13, 47),
                },
            ],
        }
        journeys_for(events_by_person, self.team)
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "TrendsQuery",
                    "dateRange": {"date_to": "2024-03-22", "date_from": "2024-03-22"},
                    "series": [
                        {
                            "kind": "EventsNode",
                            "event": None,
                            "name": "All events",
                            "math": "dau",
                        },
                        {
                            "kind": "EventsNode",
                            "event": "$pageview",
                            "name": "$pageview",
                            "math": "dau",
                        },
                    ],
                    "interval": "day",
                    "trendsFilter": {"showLegend": True, "aggregationAxisFormat": "percentage", "formula": "(B/A)*100"},
                }
            },
        )
        exported_asset.save()
        mocked_uuidt.return_value = "a-guid"

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)
            lines = (content or "").strip().split("\r\n")
            self.assertEqual(
                lines,
                ["series,22-Mar-2024", "Formula ((B/A)*100),100.0"],
            )

    def test_csv_exporter_trends_query_with_compare_previous_option(
        self,
    ) -> None:
        _create_person(distinct_ids=[f"user_1"], team=self.team)

        date = datetime(2023, 3, 21, 13, 46)
        date_next_week = date + relativedelta(days=7)

        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date,
            properties={"$browser": "Safari"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date,
            properties={"$browser": "Chrome"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date,
            properties={"$browser": "Chrome"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date,
            properties={"$browser": "Firefox"},
        )

        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date_next_week,
            properties={"$browser": "Chrome"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date_next_week,
            properties={"$browser": "Chrome"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date_next_week,
            properties={"$browser": "Chrome"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date_next_week,
            properties={"$browser": "Firefox"},
        )
        _create_event(
            event="$pageview",
            distinct_id="1",
            team=self.team,
            timestamp=date_next_week,
            properties={"$browser": "Firefox"},
        )

        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "TrendsQuery",
                    "dateRange": {
                        "date_from": date.strftime("%Y-%m-%d"),
                        "date_to": date_next_week.strftime("%Y-%m-%d"),
                    },
                    "series": [
                        {
                            "kind": "EventsNode",
                            "event": "$pageview",
                            "name": "$pageview",
                            "math": "total",
                        },
                    ],
                    "interval": "day",
                    "compareFilter": {"compare": True, "compare_to": "-1w"},
                    "breakdownFilter": {"breakdown": "$browser", "breakdown_type": "event"},
                }
            },
        )
        exported_asset.save()

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)  # type: ignore

            lines = (content or "").strip().splitlines()

            expected_lines = [
                "series,21-Mar-2023,22-Mar-2023,23-Mar-2023,24-Mar-2023,25-Mar-2023,26-Mar-2023,27-Mar-2023,28-Mar-2023",
                "Chrome - current,2.0,0.0,0.0,0.0,0.0,0.0,0.0,3.0",
                "Firefox - current,1.0,0.0,0.0,0.0,0.0,0.0,0.0,2.0",
                "Safari - current,1.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0",
                "Chrome - previous,0.0,0.0,0.0,0.0,0.0,0.0,0.0,2.0",
                "Firefox - previous,0.0,0.0,0.0,0.0,0.0,0.0,0.0,1.0",
                "Safari - previous,0.0,0.0,0.0,0.0,0.0,0.0,0.0,1.0",
            ]

            self.assertEqual(lines, expected_lines)

    def test_csv_exporter_trends_actors(
        self,
    ) -> None:
        with freeze_time("2022-06-01T12:00:00.000Z"):
            _create_person(distinct_ids=[f"user_1"], team=self.team, uuid="4beb316f-23aa-2584-66d3-4a1b8ab458f2")

        events_by_person = {
            "user_1": [
                {
                    "event": "$pageview",
                    "timestamp": datetime(2020, 3, 22, 13, 46),
                },
                {
                    "event": "$pageview",
                    "timestamp": datetime(2020, 3, 22, 13, 47),
                },
            ],
        }
        journeys_for(events_by_person, self.team)
        _create_event(
            event="$pageview",
            distinct_id="user_2",  # personless user
            person_id="d0780d6b-ccd0-44fa-a227-47efe4f3f30d",
            timestamp=datetime(2020, 3, 22, 13, 48),
            team=self.team,
        )
        flush_persons_and_events()

        exported_asset = ExportedAsset(
            team=self.team,
            export_format=ExportedAsset.ExportFormat.CSV,
            export_context={
                "source": {
                    "kind": "ActorsQuery",
                    "search": "",
                    "select": ["actor", "event_count"],
                    "source": {
                        "day": "2020-03-22T00:00:00Z",
                        "kind": "InsightActorsQuery",
                        "series": 0,
                        "source": {
                            "kind": "TrendsQuery",
                            "series": [
                                {"kind": "EventsNode", "math": "total", "name": "$pageview", "event": "$pageview"}
                            ],
                            "trendsFilter": {},
                        },
                        "includeRecordings": False,
                    },
                    "orderBy": ["event_count DESC, actor_id DESC"],
                }
            },
        )
        exported_asset.save()

        with self.settings(OBJECT_STORAGE_ENABLED=True, OBJECT_STORAGE_EXPORTS_FOLDER="Test-Exports"):
            csv_exporter.export_tabular(exported_asset)
            content = object_storage.read(exported_asset.content_location)  # type: ignore
            lines = (content or "").strip().split("\r\n")
            self.assertEqual(
                lines,
                [
                    "actor.id,actor.is_identified,actor.created_at,actor.distinct_ids.0,event_count,event_distinct_ids.0",
                    "4beb316f-23aa-2584-66d3-4a1b8ab458f2,False,2022-06-01 12:00:00+00:00,user_1,2,user_1",
                    "d0780d6b-ccd0-44fa-a227-47efe4f3f30d,,,user_2,1,user_2",
                ],
            )
