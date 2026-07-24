"""Storage contract for user-uploaded source files.

Framework-free and source-agnostic on purpose: the upload endpoint and the table-create endpoint
(both in ``data_warehouse``) share this key layout, so it has to live in exactly one place. Uploads
land in PostHog's own data warehouse bucket, namespaced by team then by a per-upload id, which is
what keeps a table's read path scoped to its own team's files.

An uploaded file becomes a self-managed ``DataWarehouseTable`` pointing straight at the stored
object: PostHog reads it in place from its own bucket, so there is no import pipeline and no
recurring sync — the same shape as a linked S3/GCS bucket, just hosted by us.
"""

import os
from typing import TYPE_CHECKING

from django.conf import settings

if TYPE_CHECKING:
    import pyarrow as pa

# Top-level bucket folder for user-uploaded files.
FILE_UPLOADS_FOLDER = "file_uploads"

# Formats a user can upload. Lowercase tokens accepted by the upload endpoint and mapped to a
# ClickHouse read format (`FILE_FORMAT_TO_TABLE_FORMAT`) when the table is created.
FORMAT_CSV = "csv"
FORMAT_JSON = "json"
FORMAT_PARQUET = "parquet"

# Excel is an upload-only input format, never a stored/read format: ClickHouse has no Excel reader,
# so the upload endpoint converts the workbook to Parquet before storing it. It therefore appears in
# UPLOAD_ACCEPTED_FORMATS (what a client may send) but not in SUPPORTED_FILE_FORMATS (what a table is
# built from) — by the time create_from_upload runs, the object is already Parquet.
FORMAT_XLSX = "xlsx"

SUPPORTED_FILE_FORMATS = (FORMAT_CSV, FORMAT_JSON, FORMAT_PARQUET)

# Formats the upload endpoint accepts as input. Superset of SUPPORTED_FILE_FORMATS because Excel is
# converted to Parquet on the way in rather than read in place.
UPLOAD_ACCEPTED_FORMATS = (*SUPPORTED_FILE_FORMATS, FORMAT_XLSX)

# Maps an uploaded file's format to the `DataWarehouseTable.TableFormat` value ClickHouse reads it
# with in place. CSV is assumed to carry a header row (the common export shape), and JSON is read as
# newline-delimited rows — the same format a self-managed S3 JSON table uses. Kept as plain strings
# so this module stays free of the model import.
FILE_FORMAT_TO_TABLE_FORMAT: dict[str, str] = {
    FORMAT_CSV: "CSVWithNames",
    FORMAT_JSON: "JSONEachRow",
    FORMAT_PARQUET: "Parquet",
}

# Cap on uploads streamed through the web pod. Larger datasets belong on a self-managed S3/GCS
# source, where PostHog reads the customer's bucket directly instead of hosting the bytes.
MAX_UPLOAD_SIZE_BYTES = 50 * 1024 * 1024


def build_file_upload_s3_prefix(team_id: int, upload_id: str) -> str:
    """Folder holding one upload's object, keyed by team then upload id."""
    return f"{FILE_UPLOADS_FOLDER}/team_{team_id}/{upload_id}"


def build_file_upload_s3_key(team_id: int, upload_id: str, filename: str) -> str:
    """Bucket-relative S3 key for one uploaded file."""
    return f"{build_file_upload_s3_prefix(team_id, upload_id)}/{filename}"


def build_file_upload_s3_path(team_id: int, upload_id: str, filename: str) -> str:
    """Bucket-qualified ``bucket/key`` path, the form s3fs takes. Use this for every read and write
    of an uploaded file so the upload endpoint and the import pipeline can't drift apart."""
    return f"{settings.DATAWAREHOUSE_BUCKET}/{build_file_upload_s3_key(team_id, upload_id, filename)}"


def build_file_upload_s3_uri(team_id: int, upload_id: str, filename: str) -> str:
    """Full ``s3://`` URI into the data warehouse bucket for one uploaded file."""
    return f"s3://{build_file_upload_s3_path(team_id, upload_id, filename)}"


def build_file_upload_url_pattern(team_id: int, upload_id: str, filename: str) -> str:
    """``https://`` URL used as the self-managed table's ``url_pattern``.

    This is the form `DataWarehouseTable.get_columns` builds its ClickHouse s3 table function from.
    The object lives in PostHog's own bucket, so the table carries no credential and reads fall back
    to the node role — never a user-supplied key. Built server-side from the source's own team, so a
    client-supplied ``upload_id`` can only ever resolve inside that team's folder.
    """
    return f"https://{settings.DATAWAREHOUSE_BUCKET_DOMAIN}/{build_file_upload_s3_key(team_id, upload_id, filename)}"


class ExcelConversionError(Exception):
    """An uploaded Excel workbook couldn't be turned into a Parquet table. The message is safe to
    surface to the user (it names what to fix), so the upload endpoint maps it straight to a 400."""


def excel_stored_filename(original_filename: str) -> str:
    """Name the converted Parquet object is stored under, derived from the uploaded workbook's name
    (``sales.xlsx`` -> ``sales.parquet``). The stored object is Parquet, so the name must match."""
    stem = os.path.splitext(original_filename)[0] or "upload"
    return f"{stem}.parquet"


def _dedupe_excel_headers(header: tuple) -> list[str]:
    """Turn a header row into unique, non-empty column names. Blank cells become ``column_N`` (1-based
    position) and repeated names get a ``_2``/``_3`` suffix, so the Parquet schema has valid, distinct
    columns for ClickHouse."""
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(header):
        name = str(value).strip() if value is not None and str(value).strip() else f"column_{index + 1}"
        count = seen.get(name, 0)
        seen[name] = count + 1
        names.append(name if count == 0 else f"{name}_{count + 1}")
    return names


def excel_to_parquet_bytes(data: bytes) -> bytes:
    """Convert an ``.xlsx``/``.xlsm`` workbook's first sheet to Parquet bytes.

    ClickHouse can't read Excel, so a self-managed table can't point at the workbook directly. We read
    the first sheet with openpyxl and re-encode it as Parquet via pyarrow — both already ship in the
    image (openpyxl powers the xlsx *export* path) and neither pulls in pandas. Cell values keep their
    native types (numbers, dates), so the Parquet schema is richer than a CSV round-trip. Only the
    first sheet is converted — a workbook maps to one table, matching the one-file-one-table contract.

    openpyxl/pyarrow are only needed on this path, so they're imported lazily to keep them off the
    module (and Django startup) import path.
    """
    import io  # noqa: PLC0415 — keep the heavy Excel/Parquet stack off the import path

    import pyarrow as pa  # noqa: PLC0415
    import pyarrow.parquet as pq  # noqa: PLC0415
    from openpyxl import load_workbook  # noqa: PLC0415

    try:
        # read_only streams rows without holding the whole sheet in memory; data_only returns the last
        # computed value of a formula cell rather than the formula text. openpyxl reads .xlsx/.xlsm
        # only, so a mis-typed .xls (which needs the xlrd engine we don't ship) fails here clearly.
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise ExcelConversionError(
            "Could not read the Excel file. Make sure it's a valid .xlsx or .xlsm workbook."
        ) from error

    try:
        rows = workbook.worksheets[0].iter_rows(values_only=True)
        header = next(rows, None)
        if not header or not any(cell is not None and str(cell).strip() for cell in header):
            raise ExcelConversionError("The first sheet has no columns. Add a header row and try again.")

        column_names = _dedupe_excel_headers(header)
        columns: list[list[object]] = [[] for _ in column_names]
        for row in rows:
            for index in range(len(column_names)):
                columns[index].append(row[index] if index < len(row) else None)
    finally:
        workbook.close()

    table = pa.table({name: _excel_column_to_arrow(values) for name, values in zip(column_names, columns)})

    buffer = io.BytesIO()
    pq.write_table(table, buffer)
    return buffer.getvalue()


def _excel_column_to_arrow(values: "list[object]") -> "pa.Array":
    """Build a pyarrow array from one column's cell values, inferring the type. A column with
    incompatible mixed types (e.g. numbers and text in the same column) can't be inferred, so it
    falls back to strings — the same shape a dataframe's object column would take."""
    import pyarrow as pa  # noqa: PLC0415 — keep pyarrow off the module import path

    try:
        return pa.array(values)
    except (pa.ArrowInvalid, pa.ArrowTypeError):
        return pa.array([None if value is None else str(value) for value in values], type=pa.string())


def hosted_upload_s3_path(url_pattern: str) -> str | None:
    """The bucket-qualified ``bucket/key`` path (the form s3fs takes) backing a self-managed table
    whose file PostHog hosts in its own data warehouse bucket, or ``None`` when the table reads from
    anywhere else — most importantly a customer-linked S3/GCS bucket, which is never ours to delete.

    The gate is the URL host: only ``url_pattern``s under ``DATAWAREHOUSE_BUCKET_DOMAIN`` are hosted
    by us. That covers both the current ``file_uploads/`` prefix and the legacy ``managed/`` one.
    """
    domain = settings.DATAWAREHOUSE_BUCKET_DOMAIN
    bucket = settings.DATAWAREHOUSE_BUCKET
    if not domain or not bucket:
        return None
    prefix = f"https://{domain}/"
    if not url_pattern.startswith(prefix):
        return None
    key = url_pattern[len(prefix) :]
    if not key:
        return None
    return f"{bucket}/{key}"
