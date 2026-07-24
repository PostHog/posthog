"""Reading an uploaded Excel workbook.

The workbook lives in PostHog's own data warehouse bucket (put there by the file-upload endpoint),
so there are no credentials to manage — the sheet is read with the node role and each sheet becomes
one warehouse table.

Reads run on the data-warehouse Temporal workers, not in a web request: openpyxl is pure Python and
costs roughly 13 seconds per million cells, which is fine on a worker and is not fine inline.
"""

import io
from collections.abc import Iterator
from typing import Any

from django.conf import settings

from openpyxl import load_workbook

from products.data_warehouse.backend.facade.api import get_s3_client
from products.warehouse_sources.backend.facade.api import build_file_upload_s3_path

# Rows buffered per yield. The pipeline batches on top of this (5000 rows / 200 MiB); the point here
# is only to avoid materializing a whole sheet before the first batch reaches it.
ROW_CHUNK = 1000


class ExcelReadError(Exception):
    """The workbook couldn't be read. The message is safe to surface to the user."""


def _uploaded_workbook_bytes(team_id: int, upload_id: str, filename: str) -> bytes:
    """Fetch the stored workbook. Read whole rather than streamed: an .xlsx is a ZIP whose reader
    seeks all over the archive, so a network file object would issue a range request per jump. The
    upload endpoint caps files at 50 MB, which is comfortable to hold on a worker."""
    path = build_file_upload_s3_path(team_id, upload_id, filename)
    try:
        with get_s3_client().open(path, "rb") as handle:
            return handle.read()
    except FileNotFoundError as error:
        raise ExcelReadError(
            "The uploaded file is no longer in storage. Upload the workbook again to re-create this source."
        ) from error


def _open_workbook(data: bytes):
    try:
        # read_only streams rows instead of building a cell object graph; data_only returns a formula
        # cell's last computed value rather than the formula text. openpyxl reads .xlsx/.xlsm only,
        # so a mis-typed legacy .xls fails here with a clear message.
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise ExcelReadError(
            "Could not read the Excel file. Make sure it's a valid .xlsx or .xlsm workbook."
        ) from error


def dedupe_headers(header: tuple) -> list[str]:
    """Turn a header row into unique, non-empty column names. Blank cells become ``column_N`` (1-based
    position) and repeats get a ``_2``/``_3`` suffix, so the resulting table has valid, distinct
    columns.

    The suffix search skips names already taken, so a header that itself contains a suffixed name
    (``A``, ``A_2``, ``A``) still yields distinct columns rather than colliding on ``A_2``.
    """
    names: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(header):
        base = str(value).strip() if value is not None and str(value).strip() else f"column_{index + 1}"
        name = base
        suffix = 2
        while name in used:
            name = f"{base}_{suffix}"
            suffix += 1
        used.add(name)
        names.append(name)
    return names


def list_sheets(team_id: int, upload_id: str, filename: str) -> list[tuple[str, list[str]]]:
    """Every sheet in the workbook as ``(sheet_name, column_names)``.

    Drives schema discovery: one sheet becomes one warehouse table, and the column names feed the
    column-selection picker. A sheet with no usable header row is skipped rather than surfaced as an
    unimportable table.
    """
    workbook = _open_workbook(_uploaded_workbook_bytes(team_id, upload_id, filename))
    try:
        sheets: list[tuple[str, list[str]]] = []
        for worksheet in workbook.worksheets:
            header = next(worksheet.iter_rows(values_only=True), None)
            if not header or not any(cell is not None and str(cell).strip() for cell in header):
                continue
            sheets.append((worksheet.title, dedupe_headers(header)))
        return sheets
    finally:
        workbook.close()


def read_sheet_rows(
    team_id: int,
    upload_id: str,
    filename: str,
    sheet_name: str,
    enabled_columns: list[str] | None = None,
) -> Iterator[list[dict[str, Any]]]:
    """Yield the sheet's rows as chunks of dicts, keyed by the deduped header names.

    Cell values keep their native Python types (numbers stay numbers, dates stay dates) so the
    pipeline infers a real schema instead of everything landing as strings. ``enabled_columns``
    projects the row down to the user's selection.
    """
    workbook = _open_workbook(_uploaded_workbook_bytes(team_id, upload_id, filename))
    try:
        if sheet_name not in workbook.sheetnames:
            raise ExcelReadError(
                f"Sheet '{sheet_name}' is no longer in the workbook. Refresh the source's schemas and try again."
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)

        header = next(rows, None)
        if not header:
            return
        columns = dedupe_headers(header)
        keep = [name in enabled_columns for name in columns] if enabled_columns is not None else None

        chunk: list[dict[str, Any]] = []
        for row in rows:
            record = {
                name: (row[index] if index < len(row) else None)
                for index, name in enumerate(columns)
                if keep is None or keep[index]
            }
            chunk.append(record)
            if len(chunk) >= ROW_CHUNK:
                yield chunk
                chunk = []
        if chunk:
            yield chunk
    finally:
        workbook.close()


def uploaded_file_exists(team_id: int, upload_id: str, filename: str) -> bool:
    """Whether the referenced upload is still in our bucket. Used to validate the source before it's
    created, so a stale reference fails at setup instead of on the first sync."""
    if not settings.DATAWAREHOUSE_BUCKET:
        return False
    return bool(get_s3_client().exists(build_file_upload_s3_path(team_id, upload_id, filename)))
